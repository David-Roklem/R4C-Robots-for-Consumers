from openpyxl import Workbook
from openpyxl.styles import Alignment

from robots.models import Robot

from datetime import datetime, timedelta


def generate_excel_report():
    '''Generate report of xlsx format using openpyxl library'''
    workbook = Workbook()

    # Establish starting and ending period - i.o. time interval for the report
    now_datetime = datetime.now()
    week_ago_datetime = now_datetime - timedelta(days=7)

    titles = ['Model', 'Version', 'Assembled for the last week (qnt.)']

    robot_models = Robot.objects.values_list('model', flat=True)
    for model in robot_models:
        # Creating worksheet with the model name and designate all the titles
        worksheet = workbook.create_sheet(model)
        worksheet.append(titles)

        # Filter out the robots assembled for the last week
        robots_assembled_last_week = Robot.objects.filter(
            model=model, created__range=(week_ago_datetime, now_datetime)
        )
        # Count the number of assembled robots of the same version
        version_count = {}
        for robot in robots_assembled_last_week:
            version = robot.version
            version_count[version] = version_count.get(version, 0) + 1

        # Fill in the table with the corresponding data
        for version, count in version_count.items():
            worksheet.append([model, version, count])

        # Align the text to the center for each cell
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=1, max_col=3
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    workbook.save('robot_report.xlsx')
